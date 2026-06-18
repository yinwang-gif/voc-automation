from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


class ExcelReporter:
    def __init__(self, config: Dict[str, Any]):
        self.output_dir = Path(config.get("output", {}).get("report_dir", ".")).expanduser()

    def generate(self, analysis_result: Dict[str, Any]) -> str:
        """生成 Excel 报告。"""
        timestamp = datetime.now().strftime("%Y%m%d")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        report_path = self.output_dir / f"VOC_分析结论表_{timestamp}.xlsx"

        df = self._build_dataframe(analysis_result)

        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="问题列表", index=False)
            self._apply_formatting(writer.sheets["问题列表"])

        return str(report_path)

    def _build_dataframe(self, result: Dict[str, Any]) -> pd.DataFrame:
        """构建数据表。"""
        rows = []
        for issue in result.get("issues", []):
            decision = issue.get("product_decision") or {}
            decision_text = (
                f"{decision.get('decision', '')}（{decision.get('decision_label', '')}）" if decision else ""
            )
            rows.append(
                {
                    "优先级": issue.get("priority", ""),
                    "问题类型": issue.get("category", ""),
                    "问题标题": issue.get("title", ""),
                    "频次": issue.get("frequency", 0),
                    "数据来源": ", ".join(issue.get("data_sources", [])),
                    "问题描述": issue.get("description", ""),
                    "建议操作": "\n".join(issue.get("suggestions", [])),
                    "产品决策": decision_text,
                    "决策理由": decision.get("rationale", ""),
                    "下一步": "\n".join(decision.get("do_next", [])),
                    "先别做": "\n".join(decision.get("do_not_yet", [])),
                }
            )
        return pd.DataFrame(
            rows,
            columns=[
                "优先级", "问题类型", "问题标题", "频次", "数据来源", "问题描述", "建议操作",
                "产品决策", "决策理由", "下一步", "先别做",
            ],
        )

    def _apply_formatting(self, worksheet) -> None:
        if worksheet.max_row < 1:
            return

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = {
            "A": 10,
            "B": 18,
            "C": 34,
            "D": 10,
            "E": 24,
            "F": 48,
            "G": 48,
            "H": 16,
            "I": 40,
            "J": 36,
            "K": 30,
        }
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width

        priority_fills = {
            "P0": PatternFill("solid", fgColor="F4CCCC"),
            "P1": PatternFill("solid", fgColor="FFF2CC"),
            "P2": PatternFill("solid", fgColor="D9EAD3"),
        }

        for row in worksheet.iter_rows(min_row=2):
            priority = row[0].value
            if priority in priority_fills:
                row[0].fill = priority_fills[priority]
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
